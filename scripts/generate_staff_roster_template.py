#!/usr/bin/env python3
"""Generate the Arabic REL-15 staff/identity intake workbook."""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


DEFAULT_OUTPUT = Path("docs/templates/REL15_STAFF_ROSTER_TEMPLATE.xlsx")
BRANCH_NAME = "El Kheima Beach Resort"
ROLES = (
    ("مدير نظام / إدارة عامة", "admin", "إدارة واسعة؛ يمنح لشخص موثوق فقط"),
    ("مدير تشغيل", "manager", "إدارة التشغيل والتقارير دون تحكم سوبر أدمن"),
    ("محاسب", "accountant", "المالية والمحاسبة؛ 2FA إلزامي"),
    ("مسؤول موارد بشرية", "hr_manager", "ملفات الموظفين والحضور والرواتب الداخلية"),
    ("مشرف", "supervisor", "إشراف تشغيلي"),
    ("استقبال", "receptionist", "الاستقبال والحجوزات والخدمة"),
    ("كاشير", "cashier", "نقطة البيع والوردية النقدية"),
    ("نادل", "waiter", "طلبات المطعم"),
    ("شيف", "chef", "تشغيل المطبخ"),
    ("موظف مطبخ", "kitchen", "شاشة المطبخ"),
    ("مدير تايم شير", "timeshare_admin", "إدارة الملكية الجزئية والتحصيل غير النقدي"),
    ("خدمة عملاء / زيارات تايم شير", "timeshare_agent", "الزيارات والدعم؛ التحصيل يحتاج استثناء مسمى"),
    ("موظف بصلاحيات محددة", "employee", "صلاحيات محددة يراجعها المسؤول"),
)


def _style_title(cell) -> None:
    cell.fill = PatternFill("solid", fgColor="173F35")
    cell.font = Font(color="FFFFFF", bold=True, size=16)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _style_header(cell) -> None:
    cell.fill = PatternFill("solid", fgColor="2F6B5B")
    cell.font = Font(color="FFFFFF", bold=True, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(bottom=Side(style="thin", color="B7CEC6"))


def build_workbook() -> Workbook:
    workbook = Workbook()
    workbook.properties.title = "REL-15 Staff Roster — El Kheima Beach Resort"
    workbook.properties.subject = "Named staff accounts and HR linkage intake"
    workbook.properties.creator = "Resort OS"

    instructions = workbook.active
    instructions.title = "اقرأ أولاً"
    instructions.sheet_view.rightToLeft = True
    instructions.merge_cells("A1:F1")
    instructions["A1"] = "طلب بيانات حسابات الموظفين — REL-15"
    _style_title(instructions["A1"])
    instructions.row_dimensions[1].height = 32
    instructions["A3"] = "الفرع التشغيلي"
    instructions["B3"] = BRANCH_NAME
    instructions["A3"].font = Font(bold=True)
    instructions["B3"].fill = PatternFill("solid", fgColor="E7F3EE")
    instructions["B3"].font = Font(bold=True, color="173F35")

    guidance = (
        "اكتب صفًا واحدًا لكل شخص فعلي؛ ممنوع حساب باسم قسم أو وردية.",
        "لا تكتب أي كلمة مرور أو كود 2FA أو كود استرداد داخل الملف.",
        "المالك والسوبر أدمن لا يُضافان هنا؛ لهما مسار أمني مستقل.",
        "لو الموظف موجود بالفعل في شاشة الموارد البشرية: اختر نعم واكتب كوده الحالي.",
        "لو الموظف جديد: اختر لا واملأ كود الموظف والمسمى وتاريخ التعيين والراتب الأساسي.",
        "الراتب هنا تشغيلي داخلي؛ لا يُعتبر اعتمادًا قانونيًا للضرائب أو التأمينات.",
        "اترك أي بيانات غير متأكد منها فارغة واكتب توضيحًا في الملاحظات؛ لا تخمّن.",
    )
    instructions["A5"] = "طريقة التعبئة"
    instructions["A5"].font = Font(bold=True, size=13, color="173F35")
    for row, item in enumerate(guidance, start=6):
        instructions[f"A{row}"] = f"{row - 5}. {item}"
        instructions.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        instructions[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        instructions.row_dimensions[row].height = 28
    instructions["A15"] = "بعد التعبئة"
    instructions["A15"].font = Font(bold=True, size=13, color="173F35")
    instructions["A16"] = (
        "أعد رفع نفس الملف. ستتم مراجعة التكرار والدور وربط HR أولًا، ثم إنشاء "
        "كلمة مرور مؤقتة آمنة لكل شخص وتسليمها منفصلة عن الملف."
    )
    instructions.merge_cells("A16:F17")
    instructions["A16"].alignment = Alignment(wrap_text=True, vertical="top")
    for column, width in {"A": 28, "B": 32, "C": 18, "D": 18, "E": 18, "F": 18}.items():
        instructions.column_dimensions[column].width = width

    roster = workbook.create_sheet("بيانات الموظفين")
    roster.sheet_view.rightToLeft = True
    roster.freeze_panes = "A5"
    roster.auto_filter.ref = "A4:P104"
    roster.merge_cells("A1:P1")
    roster["A1"] = f"موظفو {BRANCH_NAME} — صف واحد لكل شخص"
    _style_title(roster["A1"])
    roster.row_dimensions[1].height = 32
    roster.merge_cells("A2:P2")
    roster["A2"] = "ممنوع إدخال كلمات المرور أو أكواد 2FA. الحقول ذات النجمة مطلوبة."
    roster["A2"].font = Font(color="9C2B2B", bold=True)
    roster["A2"].alignment = Alignment(horizontal="center")

    headers = (
        "م *", "الاسم الكامل *", "البريد الإلكتروني للحساب *", "رقم المحمول",
        "الوظيفة / الدور *", "موجود في HR؟ *", "كود الموظف الحالي",
        "كود موظف جديد", "المسمى الوظيفي", "القسم", "تاريخ التعيين",
        "الراتب الأساسي الشهري", "وعاء التأمين (اختياري)",
        "الرقم القومي (اختياري)", "لغة الواجهة *", "ملاحظات",
    )
    widths = (7, 28, 34, 18, 32, 16, 20, 19, 24, 21, 17, 22, 22, 23, 16, 36)
    for column, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = roster.cell(4, column, header)
        _style_header(cell)
        roster.column_dimensions[get_column_letter(column)].width = width
    roster.row_dimensions[4].height = 42

    pale_green = PatternFill("solid", fgColor="F1F8F5")
    pale_yellow = PatternFill("solid", fgColor="FFF8E1")
    thin = Side(style="hair", color="D8E5E0")
    for row in range(5, 105):
        roster.cell(row, 1, row - 4)
        for column in range(1, 17):
            cell = roster.cell(row, column)
            cell.alignment = Alignment(vertical="center", wrap_text=column in {2, 5, 8, 9, 10, 16})
            cell.border = Border(bottom=thin)
            if column in {2, 3, 5, 6, 15}:
                cell.fill = pale_green
            elif column in {7, 8, 9, 11, 12}:
                cell.fill = pale_yellow
        roster.cell(row, 11).number_format = "yyyy-mm-dd"
        roster.cell(row, 12).number_format = '#,##0.00'
        roster.cell(row, 13).number_format = '#,##0.00'
        roster.cell(row, 14).number_format = "@"
        roster.cell(row, 1).protection = Protection(locked=True)
        roster.row_dimensions[row].height = 25

    roles_sheet = workbook.create_sheet("مرجع الأدوار")
    roles_sheet.sheet_view.rightToLeft = True
    roles_sheet.append(("الاختيار في الملف", "رمز النظام", "الاستخدام"))
    for cell in roles_sheet[1]:
        _style_header(cell)
    for role in ROLES:
        roles_sheet.append(role)
    roles_sheet.column_dimensions["A"].width = 36
    roles_sheet.column_dimensions["B"].width = 22
    roles_sheet.column_dimensions["C"].width = 62
    for row in roles_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    roles_sheet.freeze_panes = "A2"

    role_validation = DataValidation(
        type="list",
        formula1=f"'مرجع الأدوار'!$A$2:$A${len(ROLES) + 1}",
        allow_blank=False,
    )
    hr_validation = DataValidation(type="list", formula1='"نعم,لا"', allow_blank=False)
    language_validation = DataValidation(type="list", formula1='"العربية,English"', allow_blank=False)
    roster.add_data_validation(role_validation)
    roster.add_data_validation(hr_validation)
    roster.add_data_validation(language_validation)
    role_validation.add("E5:E104")
    hr_validation.add("F5:F104")
    language_validation.add("O5:O104")

    missing_fill = PatternFill("solid", fgColor="FCE8E6")
    for column in ("B", "C", "E", "F", "O"):
        roster.conditional_formatting.add(
            f"{column}5:{column}104",
            FormulaRule(formula=[f'AND($B5<>"",{column}5="")'], fill=missing_fill),
        )
    roster.conditional_formatting.add(
        "G5:G104",
        FormulaRule(formula=['AND($B5<>"",$F5="نعم",$G5="")'], fill=missing_fill),
    )
    for column in ("H", "I", "K", "L"):
        roster.conditional_formatting.add(
            f"{column}5:{column}104",
            FormulaRule(formula=[f'AND($B5<>"",$F5="لا",{column}5="")'], fill=missing_fill),
        )

    workbook.active = 0
    return workbook


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()
    args.output.parent.mkdir(parents=True, exist_ok=True)
    build_workbook().save(args.output)
    print(args.output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
