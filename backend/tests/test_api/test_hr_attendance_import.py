"""
tests/test_api/test_hr_attendance_import.py
wagdy.md H-07 — استيراد ملف حضور Excel (services.import_attendance_excel +
POST /hr/attendance/import-excel). نفس نمط اختبارات
timeshare.services.import_contracts_excel (tests/test_api/test_timeshare.py
TestExcelImport) — _build_workbook helper، أخطاء لكل صف بدل ما توقف الاستيراد
كله، لكن هنا upsert حقيقي (مش skip-on-duplicate) لأن AttendanceRecord عنده
مفتاح طبيعي حقيقي (employee_id + record_date).
"""
from __future__ import annotations

import io
import uuid
from datetime import date, datetime, timedelta
from decimal import Decimal

import pytest
from sqlalchemy.orm import Session

from app.modules.hr import crud, services


def _build_workbook(headers: list, rows: list[list]) -> bytes:
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def branch(db: Session):
    from app.modules.core.models import Branch
    b = Branch(name="Attendance Import Branch", name_ar="فرع استيراد الحضور",
               code=f"AIB-{uuid.uuid4().hex[:6].upper()}")
    db.add(b)
    db.commit()
    return b


def _make_employee(db, branch, code=None, name="أحمد سمير"):
    from app.modules.hr.crud import create_employee
    from app.modules.hr.schemas import EmployeeCreate
    emp = create_employee(db, EmployeeCreate(
        branch_id=branch.id,
        employee_code=code or f"EMP-{uuid.uuid4().hex[:6].upper()}",
        full_name=name, position="نادل", basic_salary=Decimal("4000"),
        hire_date=date(2023, 1, 1),
    ))
    db.commit()
    return emp


class TestImportAttendanceExcelService:

    def test_import_by_employee_code(self, db: Session, branch):
        emp = _make_employee(db, branch, code="EMP-001")
        content = _build_workbook(
            ["employee_code", 1, 2, 3],
            [["EMP-001", "p", "u", "v"]],
        )
        result = services.import_attendance_excel(db, branch.id, 2026, 6, content)
        assert result.imported == 3
        assert result.errors == []
        assert result.unmatched_employees == []

        rec1 = crud.get_attendance_for_date(db, emp.id, date(2026, 6, 1))
        assert rec1.status == "present"
        rec2 = crud.get_attendance_for_date(db, emp.id, date(2026, 6, 2))
        assert rec2.status == "absent"
        rec3 = crud.get_attendance_for_date(db, emp.id, date(2026, 6, 3))
        assert rec3.status == "leave"

    def test_import_falls_back_to_full_name_match(self, db: Session, branch):
        emp = _make_employee(db, branch, name="سارة إبراهيم")
        content = _build_workbook(
            ["الموظف", 1],
            [["سارة إبراهيم", "p"]],
        )
        result = services.import_attendance_excel(db, branch.id, 2026, 6, content)
        assert result.imported == 1
        rec = crud.get_attendance_for_date(db, emp.id, date(2026, 6, 1))
        assert rec.status == "present"

    def test_name_match_is_case_insensitive_and_trims_whitespace(self, db: Session, branch):
        emp = _make_employee(db, branch, name="محمود علي")
        content = _build_workbook(
            ["name", 1],
            [["  محمود علي  ", "p"]],
        )
        result = services.import_attendance_excel(db, branch.id, 2026, 6, content)
        assert result.imported == 1

    def test_unmatched_employee_reported_not_silently_dropped(self, db: Session, branch):
        content = _build_workbook(
            ["employee_code", 1],
            [["NOBODY-999", "p"]],
        )
        result = services.import_attendance_excel(db, branch.id, 2026, 6, content)
        assert result.imported == 0
        assert result.unmatched_employees == ["NOBODY-999"]

    def test_unknown_status_code_becomes_row_error_but_batch_continues(self, db: Session, branch):
        emp = _make_employee(db, branch, code="EMP-002")
        content = _build_workbook(
            ["employee_code", 1, 2],
            [["EMP-002", "???", "p"]],
        )
        result = services.import_attendance_excel(db, branch.id, 2026, 6, content)
        assert result.imported == 1  # يوم 2 فقط نجح
        assert len(result.errors) == 1
        assert "قيمة حالة غير معروفة" in result.errors[0]

    def test_blank_cell_skipped_not_error(self, db: Session, branch):
        """يوم مستقبلي في شهر جاري — الخلية فاضية، مش لازم تتحسب خطأ."""
        emp = _make_employee(db, branch, code="EMP-003")
        content = _build_workbook(
            ["employee_code", 1, 2],
            [["EMP-003", "p", None]],
        )
        result = services.import_attendance_excel(db, branch.id, 2026, 6, content)
        assert result.imported == 1
        assert result.errors == []

    def test_blank_identifier_row_skipped(self, db: Session, branch):
        content = _build_workbook(
            ["employee_code", 1],
            [[None, "p"], ["", "p"]],
        )
        result = services.import_attendance_excel(db, branch.id, 2026, 6, content)
        assert result.imported == 0
        assert result.unmatched_employees == []
        assert result.errors == []

    def test_reimport_upserts_instead_of_duplicating(self, db: Session, branch):
        """رفع نفس الملف تاني بعد تصحيح خانة — لازم يحدّث السجل الموجود،
        مش يرفضه كتكرار (نفس فلسفة crud.upsert_attendance)."""
        emp = _make_employee(db, branch, code="EMP-004")
        content1 = _build_workbook(["employee_code", 1], [["EMP-004", "u"]])
        services.import_attendance_excel(db, branch.id, 2026, 6, content1)
        rec = crud.get_attendance_for_date(db, emp.id, date(2026, 6, 1))
        assert rec.status == "absent"

        content2 = _build_workbook(["employee_code", 1], [["EMP-004", "p"]])
        services.import_attendance_excel(db, branch.id, 2026, 6, content2)

        db.refresh(rec)
        assert rec.status == "present"

        from app.modules.hr.models import AttendanceRecord
        count = db.query(AttendanceRecord).filter(
            AttendanceRecord.employee_id == emp.id, AttendanceRecord.record_date == date(2026, 6, 1),
        ).count()
        assert count == 1  # مفيش تكرار

    def test_date_typed_column_header_used_directly(self, db: Session, branch):
        """لو عنوان العمود تاريخ حقيقي (openpyxl لما الخلية متنسّقة كتاريخ)
        بدل رقم يوم خام — يُستخدم مباشرة، مش period_year/period_month."""
        emp = _make_employee(db, branch, code="EMP-005")
        content = _build_workbook(
            ["employee_code", date(2026, 3, 15)],
            [["EMP-005", "p"]],
        )
        # period_year/period_month هنا مختلفين عمدًا عن عنوان العمود —
        # لازم عنوان العمود (تاريخ حقيقي) هو اللي يغلب
        result = services.import_attendance_excel(db, branch.id, 2026, 6, content)
        assert result.imported == 1
        rec = crud.get_attendance_for_date(db, emp.id, date(2026, 3, 15))
        assert rec is not None
        assert rec.status == "present"

    def test_day_number_out_of_range_for_month_is_row_error(self, db: Session, branch):
        emp = _make_employee(db, branch, code="EMP-006")
        content = _build_workbook(
            ["employee_code", 31],
            [["EMP-006", "p"]],
        )
        # أبريل عنده 30 يوم بس — يوم 31 مش موجود
        result = services.import_attendance_excel(db, branch.id, 2026, 4, content)
        assert result.imported == 0
        assert len(result.errors) == 1

    def test_non_day_column_ignored_silently(self, db: Session, branch):
        """عمود مش رقم يوم ولا تاريخ (زي عمود ملاحظات) يتجاهل، مش يتحسب
        عمود يوم غلط."""
        emp = _make_employee(db, branch, code="EMP-007")
        content = _build_workbook(
            ["employee_code", "notes", 1],
            [["EMP-007", "بعض الملاحظات", "p"]],
        )
        result = services.import_attendance_excel(db, branch.id, 2026, 6, content)
        assert result.imported == 1
        rec = crud.get_attendance_for_date(db, emp.id, date(2026, 6, 1))
        assert rec.status == "present"

    def test_empty_file_raises(self, db: Session, branch):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.delete_rows(1, ws.max_row)
        buf = io.BytesIO()
        wb.save(buf)

        with pytest.raises(ValueError, match="فاضي"):
            services.import_attendance_excel(db, branch.id, 2026, 6, buf.getvalue())

    def test_no_valid_day_columns_raises(self, db: Session, branch):
        content = _build_workbook(
            ["employee_code", "notes", "extra"],
            [["EMP-001", "a", "b"]],
        )
        with pytest.raises(ValueError, match="عمود يوم صالح"):
            services.import_attendance_excel(db, branch.id, 2026, 6, content)

    def test_employee_from_different_branch_treated_as_unmatched(self, db: Session, branch):
        """كود موظف عالمي فريد (unique=True) بس تابع لفرع تاني — مايتقبلش
        كموظف صالح لملف الفرع ده."""
        from app.modules.core.models import Branch
        other_branch = Branch(name="Other Branch", name_ar="فرع تاني",
                               code=f"OB-{uuid.uuid4().hex[:6].upper()}")
        db.add(other_branch)
        db.commit()
        _make_employee(db, other_branch, code="EMP-OTHER")

        content = _build_workbook(["employee_code", 1], [["EMP-OTHER", "p"]])
        result = services.import_attendance_excel(db, branch.id, 2026, 6, content)
        assert result.imported == 0
        assert result.unmatched_employees == ["EMP-OTHER"]

    def test_multiple_employees_and_days_in_one_file(self, db: Session, branch):
        emp1 = _make_employee(db, branch, code="EMP-A")
        emp2 = _make_employee(db, branch, code="EMP-B")
        content = _build_workbook(
            ["employee_code", 1, 2, 3],
            [["EMP-A", "p", "p", "u"], ["EMP-B", "u", "p", "v"]],
        )
        result = services.import_attendance_excel(db, branch.id, 2026, 6, content)
        assert result.imported == 6
        assert result.errors == []

        from app.modules.hr.models import AttendanceRecord
        count = db.query(AttendanceRecord).filter(AttendanceRecord.branch_id == branch.id).count()
        assert count == 6


class TestImportAttendanceExcelHttp:

    def _upload(self, client, branch_id, period_year, period_month, content, headers):
        return client.post(
            "/api/v1/hr/attendance/import-excel",
            params={"branch_id": branch_id, "period_year": period_year, "period_month": period_month},
            files={"file": ("attendance.xlsx", content,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=headers,
        )

    def test_import_via_http_end_to_end(self, client, db: Session, fake_redis, manager_headers, branch):
        emp = _make_employee(db, branch, code="EMP-HTTP1")
        content = _build_workbook(["employee_code", 1, 2], [["EMP-HTTP1", "p", "u"]])

        resp = self._upload(client, branch.id, 2026, 6, content, manager_headers)
        assert resp.status_code == 200, resp.text
        body = resp.json()
        assert body["imported"] == 2
        assert body["errors"] == []
        assert body["unmatched_employees"] == []

    def test_import_requires_manager_role(self, client, db: Session, fake_redis, waiter_headers, branch):
        content = _build_workbook(["employee_code", 1], [["EMP-1", "p"]])
        resp = self._upload(client, branch.id, 2026, 6, content, waiter_headers)
        assert resp.status_code == 403

    def test_import_empty_file_returns_400(self, client, db: Session, fake_redis, manager_headers, branch):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.delete_rows(1, ws.max_row)
        buf = io.BytesIO()
        wb.save(buf)

        resp = self._upload(client, branch.id, 2026, 6, buf.getvalue(), manager_headers)
        assert resp.status_code == 400

    def test_import_reports_unmatched_employees(self, client, db: Session, fake_redis, manager_headers, branch):
        content = _build_workbook(["employee_code", 1], [["GHOST-001", "p"]])
        resp = self._upload(client, branch.id, 2026, 6, content, manager_headers)
        assert resp.status_code == 200
        assert resp.json()["unmatched_employees"] == ["GHOST-001"]
