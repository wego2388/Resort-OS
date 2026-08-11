"""
tests/test_api/test_timeshare.py
Integration tests for timeshare module.
"""
from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal

import pytest
from sqlalchemy.orm import Session

from app.modules.timeshare.schemas import (
    TimeshareContractCreate, TimeshareContractUpdate, PayInstallmentRequest,
)
from app.modules.timeshare import services, crud, models


@pytest.fixture
def branch(db: Session):
    import uuid
    from app.modules.core.models import Branch
    b = Branch(name="Test", name_ar="اختبار", code=f"TS-{uuid.uuid4().hex[:6].upper()}")
    db.add(b); db.flush()
    from app.modules.finance.models import CashierShift
    db.add(CashierShift(
        branch_id=b.id, cashier_id=1, opened_by=1, opened_at=datetime.utcnow(),
        opening_float=Decimal("0"), status="open",
    ))
    db.flush()
    return b


@pytest.fixture
def unit(db: Session, branch):
    """وحدة ملكية جزئية حقيقية (2R) متاحة — لازمة عشان create_visit يقدر يخصّص
    وحدة فعلية (allocation logic حقيقي، مش مجرد سطر تاريخ من غير حجز حقيقي)."""
    from app.modules.timeshare.models import TimeshareUnit
    u = TimeshareUnit(branch_id=branch.id, unit_number="A-101", unit_type="Studio")
    db.add(u); db.flush()
    return u


@pytest.fixture
def contract(db: Session, branch):
    make_finance_accounts(db, branch)
    data = TimeshareContractCreate(
        branch_id=branch.id,
        customer_name="أحمد محمد",
        customer_phone="01000000001",
        room_type="Studio", unit_capacity=2,
        week_number=28,
        nights_per_year=7,
        total_value=Decimal("120000"),
        down_payment=Decimal("20000"),
        installments=12,
        installment_period=1,
        first_installment_date=date(2026, 8, 1),
        partner_share_pct=Decimal("0"),
        start_date=date(2026, 7, 1),
    )
    return services.create_contract(db, data, signed_by=1)



class TestTimeshareContract:

    def test_create_generates_installments(self, db, branch, contract):
        assert contract.contract_number.startswith("TS-")
        assert len(contract.installments_list) == 12

    def test_update_contract_records_actor_and_before_after_values(self, db, contract):
        import json
        from app.modules.core.models import AuditLog

        services.update_contract(
            db,
            contract.id,
            TimeshareContractUpdate(notes="تمت المراجعة"),
            updated_by=1,
        )

        audit = (
            db.query(AuditLog)
            .filter_by(action="update_contract", entity_type="timeshare_contract", entity_id=contract.id)
            .one()
        )
        assert audit.user_id == 1
        assert json.loads(audit.old_data)["notes"] is None
        assert json.loads(audit.new_data)["notes"] == "تمت المراجعة"

    def test_installment_amounts_sum_to_remaining(self, db, contract):
        total = sum(i.amount for i in contract.installments_list)
        assert total == Decimal("100000")  # 120000 - 20000

    def test_first_installment_date_correct(self, db, contract):
        first = min(contract.installments_list, key=lambda i: i.installment_no)
        assert first.due_date == date(2026, 8, 1)

    def test_down_payment_exceeds_total_raises(self, db, branch):
        data = TimeshareContractCreate(
            branch_id=branch.id,
            customer_name="عميل",
            room_type="Studio", unit_capacity=2,
            total_value=Decimal("50000"),
            down_payment=Decimal("60000"),  # أكبر من الإجمالي
            installments=12, installment_period=1,
            first_installment_date=date(2026, 8, 1),
            partner_share_pct=Decimal("0"),
            start_date=date(2026, 7, 1),
        )
        with pytest.raises(ValueError, match="الدفعة الأولى"):
            services.create_contract(db, data, signed_by=1)

    def test_end_date_before_start_date_raises(self, db, branch):
        """قاعدة عمل حقيقية من elkheima-beach-resort: end_date يجب أن يكون
        بعد start_date — كانت ناقصة في resort-os (فقط عند التحقق من الـ schema
        لم تكن هناك مقارنة بين الحقلين)."""
        data = TimeshareContractCreate(
            branch_id=branch.id,
            customer_name="عميل",
            room_type="Studio", unit_capacity=2,
            total_value=Decimal("50000"),
            down_payment=Decimal("5000"),
            installments=12, installment_period=1,
            first_installment_date=date(2026, 8, 1),
            partner_share_pct=Decimal("0"),
            start_date=date(2026, 7, 1),
            end_date=date(2026, 6, 1),  # قبل start_date
        )
        with pytest.raises(ValueError, match="تاريخ الانتهاء"):
            services.create_contract(db, data, signed_by=1)

    def test_end_date_equal_start_date_raises(self, db, branch):
        data = TimeshareContractCreate(
            branch_id=branch.id,
            customer_name="عميل",
            room_type="Studio", unit_capacity=2,
            total_value=Decimal("50000"),
            down_payment=Decimal("5000"),
            installments=12, installment_period=1,
            first_installment_date=date(2026, 8, 1),
            partner_share_pct=Decimal("0"),
            start_date=date(2026, 7, 1),
            end_date=date(2026, 7, 1),
        )
        with pytest.raises(ValueError, match="تاريخ الانتهاء"):
            services.create_contract(db, data, signed_by=1)

    def test_end_date_after_start_date_succeeds(self, db, branch):
        data = TimeshareContractCreate(
            branch_id=branch.id,
            customer_name="عميل",
            room_type="Studio", unit_capacity=2,
            total_value=Decimal("50000"),
            down_payment=Decimal("5000"),
            installments=12, installment_period=1,
            first_installment_date=date(2026, 8, 1),
            partner_share_pct=Decimal("0"),
            start_date=date(2026, 7, 1),
            end_date=date(2026, 7, 2),
        )
        make_finance_accounts(db, branch)
        c = services.create_contract(db, data, signed_by=1)
        assert c.end_date == date(2026, 7, 2)


class TestPayInstallment:

    def test_pay_full_installment(self, db, contract):
        inst = contract.installments_list[0]
        req = PayInstallmentRequest(
            paid_amount=inst.amount,
            payment_method="cash",
            receipt_number="REC-001",
        )
        paid = services.pay_installment(db, inst.id, req, collected_by=1)
        assert paid.status == "paid"
        assert paid.paid_amount == inst.amount

    def test_partial_payment(self, db, contract):
        inst = contract.installments_list[0]
        req = PayInstallmentRequest(
            paid_amount=inst.amount / 2,
            payment_method="card",
        )
        paid = services.pay_installment(db, inst.id, req, collected_by=1)
        assert paid.status == "partial"

    def test_cash_installment_is_linked_to_open_shift_and_expected_cash(self, db, contract):
        from app.modules.finance.models import Payment
        from app.modules.finance.services import build_shift_end_report

        inst = contract.installments_list[0]
        services.pay_installment(
            db, inst.id,
            PayInstallmentRequest(paid_amount=inst.amount, payment_method="cash"),
            collected_by=1,
        )
        payment = db.query(Payment).filter_by(
            source="timeshare_installment", ref_order_id=inst.id,
        ).one()
        assert payment.shift_id is not None
        report = build_shift_end_report(db, payment.shift_id)
        assert report.expected_cash == inst.amount

    def test_card_installment_is_auditable_but_excluded_from_drawer(self, db, contract):
        from app.modules.finance.models import Payment

        inst = contract.installments_list[0]
        services.pay_installment(
            db, inst.id,
            PayInstallmentRequest(paid_amount=inst.amount, payment_method="card"),
            collected_by=1,
        )
        payment = db.query(Payment).filter_by(
            source="timeshare_installment", ref_order_id=inst.id,
        ).one()
        assert payment.cashier_id == 1
        assert payment.shift_id is None

    def test_cash_installment_without_open_shift_is_rejected_atomically(self, db, contract):
        from app.modules.finance.models import CashierShift, Payment
        from app.modules.finance.services import OpenCashierShiftRequiredError

        shift = db.query(CashierShift).filter_by(branch_id=contract.branch_id, cashier_id=1).one()
        shift.status = "closed"
        db.commit()
        inst = contract.installments_list[0]
        with pytest.raises(OpenCashierShiftRequiredError):
            services.pay_installment(
                db, inst.id,
                PayInstallmentRequest(paid_amount=inst.amount, payment_method="cash"),
                collected_by=1,
            )
        db.refresh(inst)
        assert inst.paid_amount == Decimal("0")
        assert db.query(Payment).filter_by(
            source="timeshare_installment", ref_order_id=inst.id,
        ).count() == 0

    def test_payment_method_routes_to_correct_gl_account(self, db, branch, contract):
        """⚠️ باج حقيقي اتصلح (OPS-DATA-02، Phase 7): تحصيل قسط بطريقة
        bank_transfer/card كان بيترحّل Dr.1100 (كاش) دايمًا زي أي تحصيل —
        نفس فئة الباج اللي اتصلح في leasing.services في نفس الجلسة. هنا
        نتأكد إن bank_transfer فعليًا بيقيّد 1110، مش 1100."""
        from app.modules.finance.models import Account, JournalEntry, JournalLine

        # contract fixture بالفعل بتزرع 1100/1110/1120/4600/4650 (راجع
        # make_finance_accounts) وبترحّل قيد الدفعة الأولى كاش (1100) —
        # فبنفلتر هنا على قيد القسط تحديدًا (TS-INST-) بدل كل حركة 1100
        # التاريخية على الفرع، وإلا الدفعة الأولى كانت هتلوّث cash_debit.
        inst = contract.installments_list[0]
        services.pay_installment(db, inst.id, PayInstallmentRequest(
            paid_amount=inst.amount, payment_method="bank_transfer",
        ), collected_by=1)

        bank_account = db.query(Account).filter_by(branch_id=branch.id, code="1110").first()
        cash_account = db.query(Account).filter_by(branch_id=branch.id, code="1100").first()
        inst_entry = (
            db.query(JournalEntry)
            .filter(JournalEntry.reference == f"TS-INST-{contract.contract_number}-{inst.installment_no}")
            .one()
        )
        bank_debit = sum(l.debit for l in inst_entry.lines if l.account_id == bank_account.id)
        cash_debit = sum(l.debit for l in inst_entry.lines if l.account_id == cash_account.id)
        assert bank_debit == inst.amount
        assert cash_debit == Decimal("0")

    def test_cannot_pay_already_paid(self, db, contract):
        inst = contract.installments_list[0]
        req = PayInstallmentRequest(paid_amount=inst.amount, payment_method="cash")
        services.pay_installment(db, inst.id, req, collected_by=1)
        with pytest.raises(ValueError, match="مدفوع"):
            services.pay_installment(db, inst.id, req, collected_by=1)

    def test_payment_unfreezes_booking(self, db, contract):
        # تجميد الحجز يدوياً
        contract.booking_frozen = True
        db.flush(); db.commit()

        inst = contract.installments_list[0]
        req = PayInstallmentRequest(paid_amount=inst.amount, payment_method="cash")
        services.pay_installment(db, inst.id, req, collected_by=1)
        db.refresh(contract)
        assert not contract.booking_frozen

    def test_overpayment_raises(self, db, contract):
        """باج حقيقي حقيقي: كان بيتقبل أي مبلغ من غير أي حد أقصى — كاشير يدخل
        50,000 غلط على قسط قيمته 10,000 كان بيتسجّل بصمت، بدون أي رفض أو تنبيه."""
        inst = contract.installments_list[0]
        req = PayInstallmentRequest(
            paid_amount=inst.amount + Decimal("40000"), payment_method="cash",
        )
        with pytest.raises(ValueError, match="أكبر من المتبقي"):
            services.pay_installment(db, inst.id, req, collected_by=1)
        db.refresh(inst)
        assert inst.status == "pending"
        assert inst.paid_amount == Decimal("0")

    def test_overpayment_on_partial_balance_raises(self, db, contract):
        """نفس الفحص بس بعد سداد جزئي — الحد الأقصى المسموح هو المتبقي فعليًا
        (amount - paid_amount)، مش amount الأصلي بالكامل."""
        inst = contract.installments_list[0]
        half = inst.amount / 2
        services.pay_installment(
            db, inst.id, PayInstallmentRequest(paid_amount=half, payment_method="cash"), collected_by=1,
        )
        with pytest.raises(ValueError, match="أكبر من المتبقي"):
            services.pay_installment(
                db, inst.id,
                PayInstallmentRequest(paid_amount=half + Decimal("1"), payment_method="cash"), collected_by=1,
            )

    def test_cannot_pay_installment_on_cancelled_contract(self, db, contract):
        """باج حقيقي: عقد اتلغى بالكامل، بس القسط المرتبط بيه فضل قابل
        للتحصيل عن طريق الـ API — كأن الإلغاء عمره ما حصل ماليًا."""
        services.cancel_contract(db, contract.id, Decimal("0"), cancelled_by=1)
        inst = contract.installments_list[0]
        req = PayInstallmentRequest(paid_amount=inst.amount, payment_method="cash")
        with pytest.raises(ValueError, match="ملغي"):
            services.pay_installment(db, inst.id, req, collected_by=1)

    def test_cannot_pay_installment_on_expired_contract(self, db, contract):
        services.update_contract(db, contract.id, TimeshareContractUpdate(status="expired"))
        inst = contract.installments_list[0]
        req = PayInstallmentRequest(paid_amount=inst.amount, payment_method="cash")
        with pytest.raises(ValueError, match="منتهي"):
            services.pay_installment(db, inst.id, req, collected_by=1)

    def test_pay_installment_posts_journal_entry(self, db, branch, contract):
        """باج حقيقي حرج (Finance First §5.2 — بيذكر أقساط الملكية الجزئية بالاسم):
        تحصيل قسط عمره ما كان بيرحّل أي قيد يومية خالص — بعكس الدفعة الأولى،
        يعني معظم إيراد عقد الملكية الجزئية (كل الأقساط بعد الأولى) كان غايبًا
        تمامًا عن الدفاتر المحاسبية."""
        from app.modules.finance import crud as finance_crud
        # contract fixture بالفعل بتزرع الحسابات + بترحّل قيد الدفعة الأولى
        # (strict=True) — فبنفلتر على مرجع القسط تحديدًا بدل عدّ كل القيود.

        inst = contract.installments_list[0]
        services.pay_installment(
            db, inst.id, PayInstallmentRequest(paid_amount=inst.amount, payment_method="cash"), collected_by=1,
        )

        entries, total = finance_crud.list_journal_entries(db, branch.id, source="timeshare")
        inst_entries = [e for e in entries if e.reference.startswith("TS-INST-")]
        assert len(inst_entries) == 1
        entry = inst_entries[0]
        assert sum(l.debit for l in entry.lines) == sum(l.credit for l in entry.lines) == inst.amount


class TestWaitlist:

    def test_add_to_waitlist(self, db, branch, contract):
        from app.modules.timeshare.schemas import WaitlistCreate
        data = WaitlistCreate(
            branch_id=branch.id,
            contract_id=contract.id,
            requested_start=date(2026, 8, 1),
            requested_end=date(2026, 8, 7),
        )
        entry = services.add_to_waitlist(db, data)
        assert entry.position == 1
        assert entry.status == "waiting"

    def test_invalid_dates_raises(self, db, branch, contract):
        from app.modules.timeshare.schemas import WaitlistCreate
        data = WaitlistCreate(
            branch_id=branch.id, contract_id=contract.id,
            requested_start=date(2026, 8, 7),
            requested_end=date(2026, 8, 1),  # نهاية قبل البداية
        )
        with pytest.raises(ValueError, match="تاريخ النهاية"):
            services.add_to_waitlist(db, data)


class TestSalesDashboard:
    """لوحة مبيعات فريق المبيعات — pipeline + متأخرات بأرقام تليفون."""

    def test_pipeline_counts_by_status(self, db: Session, branch, contract):
        # عقد تاني في حالة draft
        data = TimeshareContractCreate(
            branch_id=branch.id, customer_name="Draft Guy", customer_phone="01111111111",
            room_type="Studio", unit_capacity=2, nights_per_year=7, total_value=Decimal("100000"),
            down_payment=Decimal("10000"), installments=12, installment_period=1,
            first_installment_date=date(2026, 8, 1), start_date=date(2026, 7, 1),
        )
        draft = services.create_contract(db, data, signed_by=1)
        services.update_contract(db, draft.id, TimeshareContractUpdate(status="draft"))

        dash = services.get_sales_dashboard(db, branch.id)
        assert dash["pipeline"]["active"] == 1
        assert dash["pipeline"]["draft"] == 1
        assert dash["active_contracts"] == 1  # draft عقود مش جوه cs-summary الأساسي

    def test_overdue_client_has_phone_for_sales_followup(self, db: Session, branch, contract):
        from app.modules.timeshare.models import TimeshareInstallment
        inst = (
            db.query(TimeshareInstallment)
            .filter(TimeshareInstallment.contract_id == contract.id, TimeshareInstallment.status == "pending")
            .first()
        )
        inst.status = "overdue"
        db.flush()

        dash = services.get_sales_dashboard(db, branch.id)
        assert dash["overdue_contracts_count"] == 1
        assert len(dash["overdue_clients"]) == 1
        overdue = dash["overdue_clients"][0]
        assert overdue["customer_phone"] == "01000000001"
        assert overdue["overdue_amount"] > 0

    def test_no_overdue_when_all_current(self, db: Session, branch, contract):
        dash = services.get_sales_dashboard(db, branch.id)
        assert dash["overdue_contracts_count"] == 0
        assert dash["overdue_clients"] == []

    def test_expired_contracts_counted_separately_from_active(self, db: Session, branch, contract):
        services.update_contract(db, contract.id, TimeshareContractUpdate(status="expired"))
        dash = services.get_sales_dashboard(db, branch.id)
        assert dash["pipeline"]["expired"] == 1
        assert dash["expired_contracts_count"] == 1
        assert dash["active_contracts"] == 0  # ماعادش نشط


def make_finance_accounts(db, branch):
    """يزرع دليل الحسابات اللي _post_deferred_revenue_journal/
    _post_installment_payment_journal/_post_maintenance_payment_journal/
    _post_contract_cancellation_refund_journal بيدوّروا عليه: 1100/1110/1120
    (نقدية/بنك/كارت — حسب _PAYMENT_METHOD_DEBIT_ACCOUNT) و4600/4650 (إيراد
    عقود/إيراد صيانة). ⚠️ 2026-07-07: بقى 4600 (revenue) بدل 2300 (كان
    liability — إيراد ملكية جزئية عمره ما كان بيتحرّر لإيراد فعلي، راجع تعليق
    _post_deferred_revenue_journal). ⚠️ 2026-08-11 (strict=True — راجع §4):
    من غير الحسابات دي، أي عملية مالية في الموديول بترفع
    FinancialConfigurationError بدل ما تكمل بصمت من غير قيد — لازم تتزرع
    قبل أي create_contract/pay_installment/pay_maintenance_due/cancel_contract
    في التستات."""
    from app.modules.finance.models import Account
    accounts = [
        Account(branch_id=branch.id, code="1100", name="Cash", account_type="asset"),
        Account(branch_id=branch.id, code="1110", name="Bank", account_type="asset"),
        Account(branch_id=branch.id, code="1120", name="Card Clearing", account_type="asset"),
        Account(branch_id=branch.id, code="4600", name="Timeshare Revenue", account_type="revenue"),
        Account(branch_id=branch.id, code="4650", name="Timeshare Maintenance Revenue", account_type="revenue"),
    ]
    db.add_all(accounts)
    db.commit()
    return accounts[0], accounts[3]


class TestContractNotFound:

    def test_get_contract_or_404_raises(self, db):
        with pytest.raises(ValueError, match="غير موجود"):
            services.get_contract_or_404(db, 999999)

    def test_update_nonexistent_contract_raises(self, db):
        with pytest.raises(ValueError):
            services.update_contract(db, 999999, TimeshareContractUpdate(status="active"))


class TestDeferredRevenueJournalPosting:
    """Gap حقيقي مماثل تماماً لـ restaurant/cafe/beach: القيد المحاسبي لدفعة
    أول عقد ملكية جزئية (Dr Cash / Cr Deferred Revenue 2300) موجود في الكود من
    زمان بس من غير أي تغطية اختبارية خالص — 0% على _post_deferred_revenue_journal."""

    def test_create_contract_posts_balanced_journal_entry(self, db: Session, branch):
        from app.modules.finance import crud as finance_crud
        cash, revenue = make_finance_accounts(db, branch)

        data = TimeshareContractCreate(
            branch_id=branch.id, customer_name="سامي عادل", room_type="Studio", unit_capacity=2,
            total_value=Decimal("80000"), down_payment=Decimal("15000"),
            installments=10, installment_period=1,
            first_installment_date=date(2026, 8, 1),
            partner_share_pct=Decimal("0"), start_date=date(2026, 7, 1),
        )
        contract = services.create_contract(db, data, signed_by=1)

        entries, total = finance_crud.list_journal_entries(db, branch.id, source="timeshare")
        assert total == 1
        entry = entries[0]
        assert entry.source_id == contract.id
        total_debit = sum(l.debit for l in entry.lines)
        total_credit = sum(l.credit for l in entry.lines)
        assert total_debit == total_credit == Decimal("15000.00")

        db.refresh(cash); db.refresh(revenue)
        cash_line = next(l for l in entry.lines if l.account_id == cash.id)
        revenue_line = next(l for l in entry.lines if l.account_id == revenue.id)
        assert cash_line.debit == Decimal("15000.00")
        assert revenue_line.credit == Decimal("15000.00")

    def test_zero_down_payment_does_not_post_journal(self, db: Session, branch):
        """دفعة أولى صفرية (down_payment=0) مفيهاش مبلغ حقيقي يترحّل."""
        from app.modules.finance import crud as finance_crud
        make_finance_accounts(db, branch)

        data = TimeshareContractCreate(
            branch_id=branch.id, customer_name="بدون دفعة", room_type="Studio", unit_capacity=2,
            total_value=Decimal("50000"), down_payment=Decimal("0"),
            installments=10, installment_period=1,
            first_installment_date=date(2026, 8, 1),
            partner_share_pct=Decimal("0"), start_date=date(2026, 7, 1),
        )
        services.create_contract(db, data, signed_by=1)

        _, total = finance_crud.list_journal_entries(db, branch.id, source="timeshare")
        assert total == 0

    def test_missing_accounts_fails_contract_creation_atomically(self, db: Session, branch):
        """⚠️ 2026-08-11: عكس السلوك القديم تمامًا — كان فيه باج محاسبي حقيقي
        هنا (لو 1100/4600 مش موجودين، العقد كان بيتسجّل عادي بصفر أثر محاسبي
        بصمت). دلوقتي strict=True: حساب مش معرَّف للفرع لازم يفشّل إنشاء
        العقد كله — مفيش عقد، مفيش أقساط، مفيش مستحق صيانة، من غير أي حالة
        نصف-مكتملة (راجع services.create_contract's try/except db.rollback())."""
        from app.modules.finance.services import FinancialConfigurationError

        data = TimeshareContractCreate(
            branch_id=branch.id, customer_name="بدون حسابات", room_type="Studio", unit_capacity=2,
            total_value=Decimal("60000"), down_payment=Decimal("10000"),
            installments=12, installment_period=1,
            first_installment_date=date(2026, 8, 1),
            partner_share_pct=Decimal("0"), start_date=date(2026, 7, 1),
        )
        with pytest.raises(FinancialConfigurationError):
            services.create_contract(db, data, signed_by=1)

        assert crud.list_contracts(db, branch.id, None, None)[1] == 0


class TestCancelContract:

    def test_cancel_sets_status_and_refund_amount(self, db: Session, contract):
        cancelled = services.cancel_contract(db, contract.id, Decimal("5000"), cancelled_by=1)
        assert cancelled.status == "cancelled"
        assert cancelled.cancel_amount == Decimal("5000")
        assert cancelled.cancelled_at is not None
        assert cancelled.cancelled_by == 1
        assert cancelled.refund_method == "cash"

    def test_cancel_rejects_refund_above_net_collected(self, db: Session, contract, branch):
        from app.modules.core.models import AuditLog
        from app.modules.finance.models import JournalEntry

        with pytest.raises(ValueError, match="أكبر من صافي المحصل"):
            services.cancel_contract(
                db,
                contract.id,
                contract.down_payment + Decimal("0.01"),
                cancelled_by=1,
            )

        db.refresh(contract)
        assert contract.status == "active"
        assert contract.cancel_amount == Decimal("0")
        assert db.query(JournalEntry).filter(
            JournalEntry.reference == f"TS-CANCEL-{contract.contract_number}",
        ).count() == 0
        assert db.query(AuditLog).filter(
            AuditLog.entity_type == "timeshare_contract",
            AuditLog.entity_id == contract.id,
            AuditLog.action == "cancel_contract",
        ).count() == 0

    def test_generic_update_cannot_bypass_secure_cancellation(self, db: Session, contract):
        with pytest.raises(ValueError, match="إجراء إلغاء العقد"):
            services.update_contract(
                db, contract.id, TimeshareContractUpdate(status="cancelled"),
            )
        db.refresh(contract)
        assert contract.status == "active"

    def test_cancel_already_cancelled_raises(self, db: Session, contract):
        services.cancel_contract(db, contract.id, Decimal("1000"), cancelled_by=1)
        with pytest.raises(ValueError, match="ملغي"):
            services.cancel_contract(db, contract.id, Decimal("500"), cancelled_by=1)

    def test_cancel_nonexistent_contract_raises(self, db: Session):
        with pytest.raises(ValueError):
            services.cancel_contract(db, 999999, Decimal("0"), cancelled_by=1)

    def test_cancel_with_refund_posts_reversal_journal_entry(self, db: Session, contract, branch):
        """باج حقيقي اتصلح: إلغاء عقد بمبلغ استرداد (cancel_amount>0) كان
        بيسجّل الرقم على العقد نفسه بس من غير أي قيد يومية — كاش حقيقي
        بيتدفع للعميل من غير أي أثر محاسبي، والإيراد اللي اتسجّل وقت
        الدفعة الأولى كان يفضل مبالغ فيه للأبد."""
        from app.modules.finance import crud as finance_crud
        from app.modules.core.models import AuditLog
        from app.modules.finance.models import Account

        # contract fixture بالفعل بتزرع 1100/4600 (راجع make_finance_accounts).
        bank = db.query(Account).filter_by(branch_id=branch.id, code="1110").first()
        revenue = db.query(Account).filter_by(branch_id=branch.id, code="4600").first()
        services.cancel_contract(
            db,
            contract.id,
            Decimal("5000"),
            refund_method="bank_transfer",
            cancelled_by=1,
        )

        entries, total = finance_crud.list_journal_entries(db, branch.id, source="timeshare")
        cancel_entries = [e for e in entries if e.reference == f"TS-CANCEL-{contract.contract_number}"]
        assert len(cancel_entries) == 1
        lines = cancel_entries[0].lines
        debit_line = next(l for l in lines if l.debit > 0)
        credit_line = next(l for l in lines if l.credit > 0)
        assert debit_line.account_id == revenue.id
        assert debit_line.debit == Decimal("5000")
        assert credit_line.account_id == bank.id
        assert credit_line.credit == Decimal("5000")
        assert cancel_entries[0].created_by == 1

        db.refresh(contract)
        assert contract.cancelled_by == 1
        assert contract.refund_method == "bank_transfer"
        audit = db.query(AuditLog).filter(
            AuditLog.entity_type == "timeshare_contract",
            AuditLog.entity_id == contract.id,
            AuditLog.action == "cancel_contract",
        ).one()
        assert audit.user_id == 1
        assert '"refund_method": "bank_transfer"' in audit.new_data

    def test_cancel_refund_fails_atomically_when_payout_account_missing(
        self, db: Session, contract, branch,
    ):
        from app.modules.core.models import AuditLog
        from app.modules.finance.models import Account, JournalEntry
        from app.modules.finance.services import FinancialConfigurationError

        bank = db.query(Account).filter_by(branch_id=branch.id, code="1110").one()
        bank.code = "1110_DISABLED"
        db.commit()

        with pytest.raises(FinancialConfigurationError):
            services.cancel_contract(
                db,
                contract.id,
                Decimal("5000"),
                refund_method="bank_transfer",
                cancelled_by=1,
            )

        db.refresh(contract)
        assert contract.status == "active"
        assert contract.cancel_amount == Decimal("0")
        assert contract.cancelled_by is None
        assert contract.refund_method is None
        assert db.query(JournalEntry).filter(
            JournalEntry.reference == f"TS-CANCEL-{contract.contract_number}",
        ).count() == 0
        assert db.query(AuditLog).filter(
            AuditLog.entity_type == "timeshare_contract",
            AuditLog.entity_id == contract.id,
            AuditLog.action == "cancel_contract",
        ).count() == 0

    def test_cancel_with_zero_refund_posts_no_journal_entry(self, db: Session, contract, branch):
        """إلغاء بمصادرة كاملة (cancel_amount=0، مفيش كاش بيرجع للعميل)
        مفيهوش أثر محاسبي جديد — مفيش حاجة تترحّل."""
        from app.modules.finance import crud as finance_crud

        services.cancel_contract(db, contract.id, Decimal("0"), cancelled_by=1)

        entries, total = finance_crud.list_journal_entries(db, branch.id, source="timeshare")
        cancel_entries = [e for e in entries if e.reference == f"TS-CANCEL-{contract.contract_number}"]
        assert len(cancel_entries) == 0


class TestTimeshareVisit:

    def test_create_visit_computes_nights(self, db: Session, branch, contract, unit):
        from app.modules.timeshare.schemas import TimeshareVisitCreate
        data = TimeshareVisitCreate(
            branch_id=branch.id, contract_id=contract.id,
            check_in=date(2026, 8, 1), check_out=date(2026, 8, 8),
        )
        visit = services.create_visit(db, data)
        assert visit.nights == 7
        assert visit.status == "scheduled"

    def test_checkout_before_checkin_raises(self, db: Session, branch, contract):
        from app.modules.timeshare.schemas import TimeshareVisitCreate
        data = TimeshareVisitCreate(
            branch_id=branch.id, contract_id=contract.id,
            check_in=date(2026, 8, 8), check_out=date(2026, 8, 1),
        )
        with pytest.raises(ValueError, match="check_out"):
            services.create_visit(db, data)

    def test_frozen_booking_blocks_visit_creation(self, db: Session, branch, contract):
        """قاعدة أعمال حقيقية: عقد بأقساط متأخرة (booking_frozen=True) ميقدرش
        يحجز زيارة جديدة لحد ما يسدّد المتأخرات."""
        from app.modules.timeshare.schemas import TimeshareVisitCreate
        contract.booking_frozen = True
        db.commit()

        data = TimeshareVisitCreate(
            branch_id=branch.id, contract_id=contract.id,
            check_in=date(2026, 8, 1), check_out=date(2026, 8, 5),
        )
        with pytest.raises(ValueError, match="مجمَّد"):
            services.create_visit(db, data)

    def test_cannot_create_visit_for_cancelled_contract(self, db: Session, branch, contract, unit):
        """باج حقيقي: عقد اتلغى بالكامل، بس كان لسه ممكن تخصّص وحدة فعلية
        لزيارة عليه — وحدة من مخزون المنتجع بتتحجز لعقد مالياً ملغي."""
        from app.modules.timeshare.schemas import TimeshareVisitCreate
        services.cancel_contract(db, contract.id, Decimal("0"), cancelled_by=1)
        data = TimeshareVisitCreate(
            branch_id=branch.id, contract_id=contract.id,
            check_in=date(2026, 8, 1), check_out=date(2026, 8, 5),
        )
        with pytest.raises(ValueError, match="ملغي"):
            services.create_visit(db, data)

    def test_cannot_create_visit_for_expired_contract(self, db: Session, branch, contract, unit):
        from app.modules.timeshare.schemas import TimeshareVisitCreate
        services.update_contract(db, contract.id, TimeshareContractUpdate(status="expired"))
        data = TimeshareVisitCreate(
            branch_id=branch.id, contract_id=contract.id,
            check_in=date(2026, 8, 1), check_out=date(2026, 8, 5),
        )
        with pytest.raises(ValueError, match="منتهي"):
            services.create_visit(db, data)

    def test_cannot_create_visit_after_contract_end_date(self, db: Session, branch, contract, unit):
        """باج حقيقي: عميل عقده انتهت مدته كان لسه يقدر ياخد وحدة فعلية من
        مخزون المنتجع — صفر تحقق من contract.end_date قبل التخصيص."""
        from app.modules.timeshare.schemas import TimeshareVisitCreate
        contract.end_date = date(2026, 7, 31)
        db.commit()
        data = TimeshareVisitCreate(
            branch_id=branch.id, contract_id=contract.id,
            check_in=date(2026, 8, 1), check_out=date(2026, 8, 5),
        )
        with pytest.raises(ValueError, match="نهاية مدة العقد"):
            services.create_visit(db, data)

    def test_visit_within_contract_end_date_still_succeeds(self, db: Session, branch, contract, unit):
        """اختبار سلبي مكمّل: زيارة داخل مدة العقد (قبل end_date) لازم تفضل
        تنجح عادي — التحقق الجديد ميمنعش الاستخدام الطبيعي."""
        from app.modules.timeshare.schemas import TimeshareVisitCreate
        contract.end_date = date(2026, 12, 31)
        db.commit()
        data = TimeshareVisitCreate(
            branch_id=branch.id, contract_id=contract.id,
            check_in=date(2026, 8, 1), check_out=date(2026, 8, 8),
        )
        visit = services.create_visit(db, data)
        assert visit.status == "scheduled"

    def test_update_visit_status(self, db: Session, branch, contract, unit):
        from app.modules.timeshare.schemas import TimeshareVisitCreate, TimeshareVisitUpdate
        data = TimeshareVisitCreate(
            branch_id=branch.id, contract_id=contract.id,
            check_in=date(2026, 8, 1), check_out=date(2026, 8, 3),
        )
        visit = services.create_visit(db, data)
        updated = services.update_visit(db, visit.id, TimeshareVisitUpdate(status="completed"))
        assert updated.status == "completed"

    def test_update_nonexistent_visit_raises(self, db: Session):
        from app.modules.timeshare.schemas import TimeshareVisitUpdate
        with pytest.raises(ValueError):
            services.update_visit(db, 999999, TimeshareVisitUpdate(status="completed"))

    # ── Real unit allocation / double-booking prevention ──────────────

    def test_floating_contract_allocates_unit(self, db: Session, branch, contract, unit):
        """عقد عائم (بدون unit_id ثابت) — لازم يتخصّص له وحدة فعلية حقيقية
        من نفس room_type لحظة إنشاء الزيارة."""
        from app.modules.timeshare.schemas import TimeshareVisitCreate
        assert contract.unit_id is None
        data = TimeshareVisitCreate(
            branch_id=branch.id, contract_id=contract.id,
            check_in=date(2026, 8, 1), check_out=date(2026, 8, 8),
        )
        visit = services.create_visit(db, data)
        assert visit.unit_id == unit.id

    def test_no_available_unit_raises(self, db: Session, branch, contract):
        """مفيش أي وحدة من نوع 2R في الفرع — لازم يرفض بوضوح (مش ينجح
        بدون تخصيص حقيقي)."""
        from app.modules.timeshare.schemas import TimeshareVisitCreate
        data = TimeshareVisitCreate(
            branch_id=branch.id, contract_id=contract.id,
            check_in=date(2026, 8, 1), check_out=date(2026, 8, 8),
        )
        with pytest.raises(ValueError, match="لا توجد وحدة متاحة"):
            services.create_visit(db, data)

    def test_floating_contract_picks_next_unit_when_first_taken(self, db: Session, branch, contract, unit):
        """عقد عائم تاني — طالما أول وحدة اتحجزت في فترة متقاطعة، لازم
        ياخد وحدة تانية متاحة، مش يفشل ومش يستخدم نفس الوحدة المحجوزة."""
        from app.modules.timeshare.models import TimeshareUnit
        from app.modules.timeshare.schemas import TimeshareContractCreate, TimeshareVisitCreate

        unit2 = TimeshareUnit(branch_id=branch.id, unit_number="A-102", unit_type="Studio")
        db.add(unit2); db.flush()

        first_visit = services.create_visit(db, TimeshareVisitCreate(
            branch_id=branch.id, contract_id=contract.id,
            check_in=date(2026, 8, 1), check_out=date(2026, 8, 8),
        ))
        assert first_visit.unit_id == unit.id

        contract2 = services.create_contract(db, TimeshareContractCreate(
            branch_id=branch.id, customer_name="عميل ثاني", room_type="Studio", unit_capacity=2,
            total_value=Decimal("120000"), down_payment=Decimal("20000"),
            installments=12, installment_period=1,
            first_installment_date=date(2026, 8, 1),
            partner_share_pct=Decimal("0"), start_date=date(2026, 7, 1),
        ), signed_by=1)

        second_visit = services.create_visit(db, TimeshareVisitCreate(
            branch_id=branch.id, contract_id=contract2.id,
            check_in=date(2026, 8, 3), check_out=date(2026, 8, 6),  # يتقاطع مع الأول
        ))
        assert second_visit.unit_id == unit2.id

    def test_permanently_assigned_unit_rejects_overlap(self, db: Session, branch, contract, unit):
        """عقد بوحدة مخصَّصة دائمًا (contract.unit_id) — زيارة تانية متقاطعة
        على نفس الوحدة لازم تُرفض بوضوح (منع تعارض حجز حقيقي)."""
        from app.modules.timeshare.schemas import TimeshareVisitCreate

        contract.unit_id = unit.id
        db.commit()

        services.create_visit(db, TimeshareVisitCreate(
            branch_id=branch.id, contract_id=contract.id,
            check_in=date(2026, 8, 1), check_out=date(2026, 8, 8),
        ))

        with pytest.raises(ValueError, match="محجوزة بالفعل"):
            services.create_visit(db, TimeshareVisitCreate(
                branch_id=branch.id, contract_id=contract.id,
                check_in=date(2026, 8, 5), check_out=date(2026, 8, 10),  # يتقاطع
            ))

    def test_permanently_assigned_unit_non_overlapping_succeeds(self, db: Session, branch, contract, unit):
        """نفس الوحدة المخصَّصة دائمًا — لكن فترة تانية غير متقاطعة لازم تنجح
        عادي (مفيش تعارض حقيقي)."""
        from app.modules.timeshare.schemas import TimeshareVisitCreate

        contract.unit_id = unit.id
        db.commit()

        services.create_visit(db, TimeshareVisitCreate(
            branch_id=branch.id, contract_id=contract.id,
            check_in=date(2026, 8, 1), check_out=date(2026, 8, 8),
        ))
        second = services.create_visit(db, TimeshareVisitCreate(
            branch_id=branch.id, contract_id=contract.id,
            check_in=date(2026, 8, 10), check_out=date(2026, 8, 15),  # مش متقاطعة
        ))
        assert second.unit_id == unit.id


class TestExcelImport:

    def _build_workbook(self, headers: list[str], rows: list[list]) -> bytes:
        import io
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        for row in rows:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_import_valid_row_creates_contract(self, db: Session, branch):
        headers = [
            "customer_name", "room_type", "total_value", "down_payment",
            "installments", "start_date", "first_installment_date",
        ]
        rows = [["ياسمين علي", "Studio", 90000, 10000, 10, "2026-07-01", "2026-08-01"]]
        content = self._build_workbook(headers, rows)

        make_finance_accounts(db, branch)
        result = services.import_contracts_excel(db, branch.id, content, signed_by=1)
        assert result["imported"] == 1
        assert result["skipped"] == 0
        assert result["errors"] == []

    def test_import_missing_required_columns_raises(self, db: Session, branch):
        headers = ["customer_name", "room_type"]  # ناقص أعمدة إلزامية
        rows = [["عميل", "Studio"]]
        content = self._build_workbook(headers, rows)

        with pytest.raises(ValueError, match="أعمدة إلزامية ناقصة"):
            services.import_contracts_excel(db, branch.id, content, signed_by=1)

    def test_import_empty_file_raises(self, db: Session, branch):
        import io
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.delete_rows(1, ws.max_row)  # لا صفوف خالص، حتى الـ header
        buf = io.BytesIO()
        wb.save(buf)

        with pytest.raises(ValueError, match="فاضي"):
            services.import_contracts_excel(db, branch.id, buf.getvalue(), signed_by=1)

    def test_import_skips_duplicate_form_number(self, db: Session, branch):
        headers = [
            "customer_name", "room_type", "total_value", "down_payment",
            "installments", "start_date", "first_installment_date", "form_number",
        ]
        rows = [
            ["عميل واحد", "Studio", 50000, 5000, 6, "2026-07-01", "2026-08-01", "FORM-100"],
            ["عميل نفس الفورم", "Studio", 60000, 6000, 6, "2026-07-01", "2026-08-01", "FORM-100"],
        ]
        content = self._build_workbook(headers, rows)

        make_finance_accounts(db, branch)
        result = services.import_contracts_excel(db, branch.id, content, signed_by=1)
        assert result["imported"] == 1
        assert result["skipped"] == 1

    def test_import_skips_duplicate_row_with_blank_form_number(self, db: Session, branch):
        """باج حقيقي اتصلح: form_number فاضي كان بيتخطى فحص التكرار بالكامل
        (الشرط كان ``if form_number and ...``) — رفع نفس الملف مرتين (شائع
        في الملفات القديمة اللي مالهاش رقم فورمة خالص) كان بيضاعف كل عقد."""
        headers = [
            "customer_name", "room_type", "total_value", "down_payment",
            "installments", "start_date", "first_installment_date",
        ]
        rows = [["ياسمين علي", "Studio", 90000, 10000, 10, "2026-07-01", "2026-08-01"]]
        content = self._build_workbook(headers, rows)

        make_finance_accounts(db, branch)
        first = services.import_contracts_excel(db, branch.id, content, signed_by=1)
        assert first["imported"] == 1
        assert first["skipped"] == 0

        second = services.import_contracts_excel(db, branch.id, content, signed_by=1)
        assert second["imported"] == 0
        assert second["skipped"] == 1

        contracts = db.query(models.TimeshareContract).filter(
            models.TimeshareContract.branch_id == branch.id,
            models.TimeshareContract.customer_name == "ياسمين علي",
        ).all()
        assert len(contracts) == 1

    def test_import_row_error_does_not_abort_whole_batch(self, db: Session, branch):
        """صف بقيمة فاسدة (down_payment أكبر من total_value) يتسجّل كـ error
        من غير ما يوقف استيراد باقي الصفوف الصحيحة."""
        headers = [
            "customer_name", "room_type", "total_value", "down_payment",
            "installments", "start_date", "first_installment_date",
        ]
        rows = [
            ["عميل فاسد", "Studio", 10000, 90000, 6, "2026-07-01", "2026-08-01"],  # down_payment > total
            ["عميل صحيح", "Studio", 50000, 5000, 6, "2026-07-01", "2026-08-01"],
        ]
        content = self._build_workbook(headers, rows)

        make_finance_accounts(db, branch)
        result = services.import_contracts_excel(db, branch.id, content, signed_by=1)
        assert result["imported"] == 1
        assert len(result["errors"]) == 1


class TestTimeshareReports:
    """تقارير الملكية الجزئية (calendar/upcoming-visits/stats/list-installments) —
    0% تغطية قبل كده رغم إنها بتستخدم فعلياً في CS/Sales dashboards."""

    def test_get_calendar_includes_booked_week(self, db: Session, branch, contract):
        # contract fixture: week_number=28
        calendar = services.get_calendar(db, branch.id, year=2026)
        assert calendar["total_booked_weeks"] == 1
        week_28_entries = [
            wk for month in calendar["calendar"] for wk in month["weeks"] if wk["week"] == 28
        ]
        assert len(week_28_entries) == 1
        assert len(week_28_entries[0]["contracts"]) == 1
        assert week_28_entries[0]["contracts"][0]["contract_number"] == contract.contract_number

    def test_get_calendar_empty_when_no_week_number(self, db: Session, branch):
        data = TimeshareContractCreate(
            branch_id=branch.id, customer_name="بدون أسبوع", room_type="Studio", unit_capacity=2,
            total_value=Decimal("40000"), down_payment=Decimal("4000"),
            installments=6, installment_period=1,
            first_installment_date=date(2026, 8, 1), start_date=date(2026, 7, 1),
        )
        make_finance_accounts(db, branch)
        services.create_contract(db, data, signed_by=1)
        calendar = services.get_calendar(db, branch.id, year=2026)
        assert calendar["total_booked_weeks"] == 0

    def test_get_upcoming_visits_finds_active_contract_within_window(self, db: Session, branch, contract):
        from datetime import date as _date
        # week_number=28 و nights_per_year=7 — احسب نافذة الزيارة القادمة يدوياً
        # مش مهم القيمة الدقيقة، المهم إن العقد النشط يظهر أو مايظهرش حسب days_until
        visits = services.get_upcoming_visits(db, branch.id, days=365)
        # نافذة أسبوع 28 خلال آخر سنة لازم تظهر ضمن نطاق الـ 365 يوم
        assert any(v["contract_number"] == contract.contract_number for v in visits)

    def test_get_upcoming_visits_excludes_non_active_contracts(self, db: Session, branch, contract):
        services.update_contract(db, contract.id, TimeshareContractUpdate(status="suspended"))
        visits = services.get_upcoming_visits(db, branch.id, days=365)
        assert not any(v["contract_number"] == contract.contract_number for v in visits)

    def test_get_stats_reflects_collected_installment(self, db: Session, branch, contract):
        inst = contract.installments_list[0]
        req = PayInstallmentRequest(paid_amount=inst.amount, payment_method="cash")
        services.pay_installment(db, inst.id, req, collected_by=1)

        stats = services.get_stats(db, branch.id)
        assert stats["collection"]["collected"] >= float(inst.amount)
        assert stats["collection"]["rate"] > 0
        assert any(r["room_type"] == "Studio" for r in stats["by_room_type"])

    def test_get_stats_by_partner_includes_resort_net_share(self, db: Session, branch):
        """صافي حصة المنتجع بعد نصيب الشريك (resort_share) — خاصية حقيقية من
        elkheima-beach-resort (khayma_share) كانت محسوبة في الـ engine
        (calculate_partner_share) لكن غير مُستخدَمة في أي مكان."""
        data = TimeshareContractCreate(
            branch_id=branch.id, customer_name="عميل شريك", room_type="Chalet", unit_capacity=4,
            total_value=Decimal("200000"), down_payment=Decimal("40000"),
            installments=10, installment_period=1,
            first_installment_date=date(2026, 8, 1), start_date=date(2026, 7, 1),
            partner_share_pct=Decimal("30"), partner_company="شركة الشريك",
        )
        make_finance_accounts(db, branch)
        services.create_contract(db, data, signed_by=1)

        stats = services.get_stats(db, branch.id)
        row = next(r for r in stats["by_partner"] if r["partner_company"] == "شركة الشريك")
        assert row["total_down"] == 40000.0
        # 40000 * (1 - 30/100) = 28000
        assert row["resort_share"] == 28000.0

    def test_list_installments_returns_summary(self, db: Session, branch, contract):
        result = services.list_installments(db, branch.id)
        assert result["total"] == 12
        assert "summary" in result

    def test_list_installments_filters_by_status(self, db: Session, branch, contract):
        inst = contract.installments_list[0]
        services.pay_installment(
            db, inst.id, PayInstallmentRequest(paid_amount=inst.amount, payment_method="cash"), collected_by=1,
        )
        result = services.list_installments(db, branch.id, status="paid")
        assert result["total"] == 1
