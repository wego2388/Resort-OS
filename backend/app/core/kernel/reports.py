"""
app/core/kernel/reports.py
ReportBuilder — PDF (table + receipt + thermal receipt) and Excel
multi-sheet report generator, owned by resort-os.

Usage:
    from app.core.kernel.reports import ReportBuilder

    rb = ReportBuilder(app_name="الخيمة بيتش ريزورت", primary_color="#1A1A2E")

    pdf_bytes = rb.table_pdf(
        title="تقرير الحجوزات اليومي",
        headers=["رقم الحجز", "الاسم", "الغرفة", "الإجمالي"],
        rows=[["BK-001", "أحمد محمد", "101", "500 EGP"]],
    )
"""

from __future__ import annotations

import os
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Optional


# ── Theme defaults ────────────────────────────────────────────────────────────

_DEFAULT_PRIMARY = "#1A1A2E"
_DEFAULT_ACCENT = "#C9A84C"
_DEFAULT_LIGHT = "#F8F9FA"
_GRAY_TEXT = "#6B7280"
_BORDER_COLOR = "#DDDDDD"

# ── Arabic-capable fonts ──────────────────────────────────────────────────────
# ⚠️ باج حقيقي اتصلح هنا (2026-08-04، طلب Mohamed بعد ما شاف إيصال حقيقي
# كل نصه العربي مربعات سودة فاضية): _t() تحت كانت بتعمل reshape+bidi صح
# للنص المنطقي، لكن الرسم الفعلي كان بيستخدم "Helvetica"/"Helvetica-Bold"
# (خط PDF قياسي Type1 بترميز WinAnsi بس) اللي مالوش أي حرف عربي واحد فيه.
#
# الحل: خط TTF عربي حقيقي (Noto Sans Arabic، مرفق في app/assets/fonts/ —
# مش خط نظام، عشان يشتغل جوه Docker container نضيف زي الإنتاج بالظبط).
#
# ⚠️ باج ثانٍ اتكشف أثناء الاختبار الحي (مش نظري): نسخة Noto Sans Arabic
# المتاحة (حزمة Debian fonts-noto-core) مقسّمة بالسكريبت فعليًا — بتغطي
# العربي + الأرقام + الترقيم الأساسي بس، من غير أي حرف لاتيني خالص (A-Z/
# a-z كلهم غير موجودين في الخط). يعني أي كلمة إنجليزية جوه نص عربي
# (Apple Juice, ORD-...) كانت هتختفي تمامًا (مش مربعات، مختفية) لو
# استخدمنا الخط ده لوحده لكل حاجة. الحل: _draw_mixed() تحت بتقسّم أي
# سطر (بعد reshape+bidi) لأجزاء عربي/غير عربي متتالية وترسم كل جزء
# بالخط اللي فعليًا بيغطي حروفه — Noto Sans Arabic للعربي، Helvetica
# (مدمج في PDF، تغطية لاتينية كاملة) لغير العربي. النتيجة: نص عربي
# وإنجليزي مختلط في نفس السطر بيتـرسم صح الاتنين، مش استبدال خط بخط.

_ASSETS_DIR = Path(__file__).resolve().parent.parent.parent / "assets"
_FONT_DIR = _ASSETS_DIR / "fonts"
_DEFAULT_LOGO_PATH = str(_ASSETS_DIR / "logo.png")

ARABIC_FONT_AVAILABLE = False
FONT_AR = "Helvetica"          # نص عربي (يقع رجوعًا لـHelvetica لو الخط مش موجود)
FONT_AR_BOLD = "Helvetica-Bold"
FONT_LATIN = "Helvetica"       # نص لاتيني/أرقام — Helvetica دايمًا كافي (base-14)
FONT_LATIN_BOLD = "Helvetica-Bold"


def _register_arabic_fonts() -> None:
    global ARABIC_FONT_AVAILABLE, FONT_AR, FONT_AR_BOLD
    reg_path = _FONT_DIR / "NotoSansArabic-Regular.ttf"
    bold_path = _FONT_DIR / "NotoSansArabic-Bold.ttf"
    if not (reg_path.is_file() and bold_path.is_file()):
        return
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        pdfmetrics.registerFont(TTFont("NotoSansArabic", str(reg_path)))
        pdfmetrics.registerFont(TTFont("NotoSansArabic-Bold", str(bold_path)))
        FONT_AR = "NotoSansArabic"
        FONT_AR_BOLD = "NotoSansArabic-Bold"
        ARABIC_FONT_AVAILABLE = True
    except Exception:
        pass


_register_arabic_fonts()


def _is_arabic_char(ch: str) -> bool:
    cp = ord(ch)
    return (
        0x0600 <= cp <= 0x06FF    # Arabic
        or 0x0750 <= cp <= 0x077F  # Arabic Supplement
        or 0x08A0 <= cp <= 0x08FF  # Arabic Extended-A
        or 0xFB50 <= cp <= 0xFDFF  # Arabic Presentation Forms-A
        or 0xFE70 <= cp <= 0xFEFF  # Arabic Presentation Forms-B
    )


def _split_script_runs(text: str) -> list[tuple[str, bool]]:
    """يقسّم نص (بعد reshape+bidi، يعني بترتيب العرض البصري النهائي) لأجزاء
    متتالية (عربي / مش عربي) — كل جزء يترسم بعدها بخط قادر فعليًا يرسم
    حروفه. أرقام/مسافات/ترقيم شائع بتفضل ملتصقة بالجزء الحالي (متاحة في
    الخطين الاتنين) عشان مايحصلش تقطيع زيادة عن اللزوم في نص زي "150.00"
    أو "2026-08-04"."""
    if not text:
        return []
    neutral = set(" \t-.,:%×/()#0123456789")
    runs: list[tuple[str, bool]] = []
    current = text[0]
    current_is_ar = _is_arabic_char(text[0])
    for ch in text[1:]:
        if ch in neutral:
            current += ch
            continue
        is_ar = _is_arabic_char(ch)
        if is_ar == current_is_ar:
            current += ch
        else:
            runs.append((current, current_is_ar))
            current = ch
            current_is_ar = is_ar
    runs.append((current, current_is_ar))
    return runs


class ReportBuilder:
    """
    Per-project report builder. Instantiate once at app startup.

    Args:
        app_name:      Shown in PDF headers/footers and Excel titles.
        primary_color: Hex color for headers (default dark navy).
        accent_color:  Hex color for totals/highlights (default gold).
        logo_path:     Absolute path to a PNG/JPG logo (optional — defaults
                       to app/assets/logo.png if present).
        rtl:           Right-to-left layout for Arabic reports (default True).
    """

    def __init__(
        self,
        app_name: str = "Resort OS",
        primary_color: str = _DEFAULT_PRIMARY,
        accent_color: str = _DEFAULT_ACCENT,
        logo_path: str = "",
        rtl: bool = True,
    ):
        self.app_name = app_name
        self.primary_color = primary_color.lstrip("#")
        self.accent_color = accent_color.lstrip("#")
        self.logo_path = logo_path or (_DEFAULT_LOGO_PATH if os.path.isfile(_DEFAULT_LOGO_PATH) else "")
        self.rtl = rtl

    # ── PDF: Table Report ─────────────────────────────────────────────────

    def table_pdf(
        self,
        title: str,
        headers: list[str],
        rows: list[list],
        *,
        subtitle: str = "",
        summary: Optional[list[tuple[str, str]]] = None,
        footer: str = "",
        landscape: bool = False,
        col_widths: Optional[list[float]] = None,   # in cm
    ) -> bytes:
        """Generate a professional table-based PDF report."""
        try:
            from reportlab.pdfgen import canvas as rl_canvas
            from reportlab.lib.pagesizes import A4, landscape as rl_landscape
            from reportlab.lib import colors
            from reportlab.lib.units import cm
        except ImportError:
            raise RuntimeError("pip install reportlab to use PDF generation")

        pagesize = rl_landscape(A4) if landscape else A4
        buf = BytesIO()
        c = rl_canvas.Canvas(buf, pagesize=pagesize)
        W, H = pagesize

        primary = colors.HexColor(f"#{self.primary_color}")
        accent = colors.HexColor(f"#{self.accent_color}")
        light = colors.HexColor(_DEFAULT_LIGHT)
        gray = colors.HexColor(_GRAY_TEXT)

        margin = 2 * cm
        usable = W - 2 * margin

        # ── Header band ──────────────────────────────────────────────────
        c.setFillColor(primary)
        c.rect(0, H - 75, W, 75, fill=True, stroke=False)

        text_start_x = margin
        if self.logo_path and os.path.isfile(self.logo_path):
            try:
                from reportlab.lib.utils import ImageReader
                c.drawImage(ImageReader(self.logo_path), margin, H - 65, 50, 50,
                            preserveAspectRatio=True, mask="auto")
                text_start_x = margin + 60
            except Exception:
                pass

        self._draw_mixed(c, text_start_x, H - 25, self.app_name, 11, bold=True,
                          color=colors.HexColor(f"#{self.accent_color}"))
        self._draw_mixed(c, text_start_x, H - 48, title, 16, bold=True, color=colors.white)

        if subtitle:
            self._draw_mixed(c, text_start_x, H - 62, subtitle, 9, color=colors.HexColor("#AAAAAA"))

        c.setFont(FONT_LATIN, 8)
        c.setFillColor(colors.HexColor("#AAAAAA"))
        ts = datetime.now().strftime("%Y-%m-%d %H:%M")
        c.drawRightString(W - margin, H - 20, ts)

        # ── Summary box ──────────────────────────────────────────────────
        y = H - 95
        if summary:
            box_h = 18
            cols = min(len(summary), 4)
            box_w = usable / cols
            c.setFillColor(light)
            c.setStrokeColor(colors.HexColor(_BORDER_COLOR))
            c.setLineWidth(0.5)
            c.roundRect(margin, y - box_h - 6, usable, box_h + 12, 4, fill=True, stroke=True)
            for i, (label, val) in enumerate(summary[:4]):
                bx = margin + i * box_w + 8
                self._draw_mixed(c, bx, y - 2, label.upper(), 7, color=gray)
                self._draw_mixed(c, bx, y - box_h + 2, str(val), 11, bold=True, color=primary)
            y -= box_h + 22

        # ── Table ─────────────────────────────────────────────────────────
        n_cols = len(headers)
        if col_widths:
            widths = [w * cm for w in col_widths]
        else:
            widths = [usable / n_cols] * n_cols

        row_h = 18
        head_h = 20

        def _draw_head_row(yy):
            xx = margin
            c.setFillColor(primary)
            c.rect(margin, yy - head_h, usable, head_h, fill=True, stroke=False)
            for hdr, w in zip(headers, widths):
                self._draw_mixed(c, xx + 4, yy - head_h + 6, str(hdr), 9, bold=True, color=colors.white)
                xx += w

        _draw_head_row(y)
        y -= head_h

        for ri, row in enumerate(rows):
            if y - row_h < 60:  # new page
                self._add_footer(c, W, footer or self.app_name, accent)
                c.showPage()
                y = H - 40
                _draw_head_row(y)
                y -= head_h

            row_bg = light if ri % 2 == 0 else colors.white
            c.setFillColor(row_bg)
            c.rect(margin, y - row_h, usable, row_h, fill=True, stroke=False)

            c.setStrokeColor(colors.HexColor(_BORDER_COLOR))
            c.setLineWidth(0.3)
            c.line(margin, y - row_h, margin + usable, y - row_h)

            x = margin
            for val, w in zip(row, widths):
                self._draw_mixed(c, x + 4, y - row_h + 5,
                                  str(val) if val is not None else "—", 9, color=colors.black)
                x += w
            y -= row_h

        self._add_footer(c, W, footer or self.app_name, accent)
        c.save()
        return buf.getvalue()

    # ── PDF: Receipt ──────────────────────────────────────────────────────

    def receipt_pdf(
        self,
        reference: str,
        title: str,
        fields: list[tuple[str, str]],
        total: float,
        currency: str = "EGP",
        *,
        note: str = "",
        footer: str = "",
        qr_data: str = "",
    ) -> bytes:
        """Generate a professional receipt / invoice PDF."""
        try:
            from reportlab.pdfgen import canvas as rl_canvas
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.units import cm
        except ImportError:
            raise RuntimeError("pip install reportlab to use PDF generation")

        buf = BytesIO()
        c = rl_canvas.Canvas(buf, pagesize=A4)
        W, H = A4
        margin = 2 * cm
        primary = colors.HexColor(f"#{self.primary_color}")
        accent = colors.HexColor(f"#{self.accent_color}")

        c.setFillColor(primary)
        c.rect(0, H - 85, W, 85, fill=True, stroke=False)

        text_start_x = margin
        if self.logo_path and os.path.isfile(self.logo_path):
            try:
                from reportlab.lib.utils import ImageReader
                c.drawImage(ImageReader(self.logo_path), margin, H - 78, 55, 55,
                            preserveAspectRatio=True, mask="auto")
                text_start_x = margin + 65
            except Exception:
                pass

        self._draw_mixed(c, text_start_x, H - 30, self.app_name, 13, bold=True,
                          color=colors.HexColor(f"#{self.accent_color}"))
        self._draw_mixed(c, text_start_x, H - 55, title, 18, bold=True, color=colors.white)

        c.setFont(FONT_LATIN, 9)
        c.setFillColor(colors.HexColor("#AAAAAA"))
        c.drawString(text_start_x, H - 70, datetime.now().strftime("%Y-%m-%d %H:%M"))

        y = H - 108
        c.setFillColor(colors.HexColor("#EEF2FF"))
        c.setStrokeColor(accent)
        c.setLineWidth(1.5)
        c.roundRect(margin, y - 10, W - 2 * margin, 30, 5, fill=True, stroke=True)
        self._draw_mixed(c, margin + 8, y + 8, f"# {reference}", 13, bold=True, color=primary)

        y -= 30
        c.setStrokeColor(colors.HexColor(_BORDER_COLOR))
        c.setLineWidth(0.5)
        for label, val in fields:
            y -= 22
            self._draw_mixed(c, margin, y, label + ":", 9, bold=True, color=colors.HexColor(_GRAY_TEXT))
            self._draw_mixed(c, margin + 5.5 * cm, y, str(val), 10, color=colors.black)
            c.line(margin, y - 4, W - margin, y - 4)

        y -= 28
        c.setFillColor(colors.HexColor("#F0F9FF"))
        c.setStrokeColor(accent)
        c.setLineWidth(2)
        c.roundRect(margin, y - 18, W - 2 * margin, 50, 8, fill=True, stroke=True)
        self._draw_mixed(c, margin + 10, y + 20, "الإجمالي / Total", 11, bold=True, color=primary)
        c.setFont(FONT_LATIN_BOLD, 22)
        c.setFillColor(accent)
        c.drawRightString(W - margin - 10, y + 12, f"{total:,.2f} {currency}")

        if note:
            self._draw_mixed(c, margin, y - 26, note, 8, color=colors.HexColor(_GRAY_TEXT))

        if qr_data:
            try:
                import qrcode as _qrcode
                qr_img = _qrcode.make(qr_data)
                qr_buf = BytesIO()
                qr_img.save(qr_buf, format="PNG")
                qr_buf.seek(0)
                from reportlab.lib.utils import ImageReader
                qr_size = 3.5 * cm
                c.drawImage(ImageReader(qr_buf),
                            W - margin - qr_size, H - 85 - qr_size - 10,
                            qr_size, qr_size)
            except ImportError:
                pass

        self._add_footer(c, W, footer or self.app_name, accent)
        c.save()
        return buf.getvalue()

    # ── PDF: Thermal Roll Receipt ───────────────────────────────────────

    def receipt_pdf_thermal(
        self,
        reference: str,
        title: str,
        fields: list[tuple[str, str]],
        total: float,
        currency: str = "EGP",
        *,
        items: Optional[list[tuple]] = None,
        summary: Optional[list[tuple[str, str]]] = None,
        subtitle: str = "",
        width_mm: float = 80.0,
        note: str = "",
        footer: str = "",
        qr_data: str = "",
        show_logo: bool = True,
    ) -> bytes:
        """Receipt PDF sized for thermal roll printers (80mm/58mm rolls).

        ``fields`` هي بيانات وصفية بسيطة بس (رقم الطلب، النوع، الطاولة...).
        ``items`` (اختياري) — قائمة (اسم، كمية، سعر الوحدة، إجمالي السطر)
        بتترسم كجدول أصناف حقيقي منفصل بصريًا عن البيانات الوصفية، بدل ما
        تتلخبط كلها في نفس عمود label:value. ``summary`` (اختياري) —
        الإجمالي قبل الضريبة/الضريبة/الخدمة/الخصم، مفصولين بصريًا عن
        الأصناف. الاتنين لو اتسابوا None، السلوك زي الإصدار القديم بالظبط
        (توافقية خلفية لباقي الموديولات — beach/leasing/timeshare/hr)."""
        try:
            from reportlab.pdfgen import canvas as rl_canvas
            from reportlab.lib import colors
            from reportlab.lib.units import mm
        except ImportError:
            raise RuntimeError("pip install reportlab to use PDF generation")

        W = width_mm * mm
        margin = 4 * mm
        line_h = 5.2 * mm
        primary = colors.HexColor(f"#{self.primary_color}")
        accent = colors.HexColor(f"#{self.accent_color}")
        gray = colors.HexColor(_GRAY_TEXT)

        has_logo = show_logo and bool(self.logo_path) and os.path.isfile(self.logo_path)
        logo_h = 15 * mm if has_logo else 0

        qr_size = min(W * 0.5, 28 * mm) if qr_data else 0
        item_lines = len(items) if items else 0
        summary_lines = len(summary) if summary else 0
        field_lines = len(fields)

        base_units = (
            8.0
            + (1.6 if subtitle else 0)
            + field_lines
            + (1.4 + item_lines * 1.8 if items else 0)
            + (0.6 + summary_lines * 0.85 if summary else 0)
            + (1 if note else 0)
        )
        H = (
            2 * margin
            + logo_h + (3 * mm if has_logo else 0)
            + base_units * line_h
            + (qr_size + 4 * mm if qr_data else 0)
            + 16 * mm
        )

        buf = BytesIO()
        c = rl_canvas.Canvas(buf, pagesize=(W, H))
        y = H - margin

        # ── Header: logo + brand ─────────────────────────────────────────
        if has_logo:
            try:
                from reportlab.lib.utils import ImageReader
                c.drawImage(ImageReader(self.logo_path), (W - logo_h) / 2, y - logo_h,
                            logo_h, logo_h, preserveAspectRatio=True, mask="auto")
            except Exception:
                has_logo = False
            y -= logo_h + 3 * mm

        self._draw_mixed(c, W / 2, y, self.app_name, 11, bold=True, color=primary, align="center")
        y -= line_h

        if subtitle:
            self._draw_mixed(c, W / 2, y, subtitle, 7, color=gray, align="center")
            y -= line_h * 0.9

        y -= line_h * 0.3
        c.setStrokeColor(accent)
        c.setLineWidth(1)
        c.line(W * 0.28, y, W * 0.72, y)
        y -= line_h * 1.1

        # ── Title + reference + timestamp ────────────────────────────────
        self._draw_mixed(c, W / 2, y, title, 10.5, bold=True, color=colors.black, align="center")
        y -= line_h * 0.95

        c.setFont(FONT_LATIN, 7)
        c.setFillColor(gray)
        c.drawCentredString(W / 2, y, datetime.now().strftime("%Y-%m-%d %H:%M"))
        y -= line_h * 1.15

        self._draw_mixed(c, W / 2, y, f"# {reference}", 9, bold=True, color=primary, align="center")
        y -= line_h * 1.2

        # ── Meta fields ───────────────────────────────────────────────────
        if fields:
            y = self._dashed_rule(c, margin, W - margin, y)
            for label, val in fields:
                self._draw_mixed(c, margin, y, label + ":", 7.3, color=gray, align="left")
                self._draw_mixed(c, W - margin, y, str(val), 7.3, bold=True, color=colors.black, align="right")
                y -= line_h

        # ── Items ─────────────────────────────────────────────────────────
        if items:
            y = self._dashed_rule(c, margin, W - margin, y)
            y -= line_h * 0.2
            self._draw_mixed(c, margin, y, "الصنف", 6.5, color=gray, align="left")
            self._draw_mixed(c, W - margin, y, "الإجمالي", 6.5, color=gray, align="right")
            y -= line_h * 1.05
            for row in items:
                name, qty, unit_price, line_total = row
                self._draw_mixed(c, margin, y, f"{qty}× {name}", 7.5, color=colors.black, align="left")
                c.setFont(FONT_LATIN_BOLD, 7.5)
                c.setFillColor(colors.black)
                c.drawRightString(W - margin, y, f"{line_total:,.2f}")
                y -= line_h * 0.85
                c.setFont(FONT_LATIN, 6.3)
                c.setFillColor(gray)
                c.drawString(margin + 3 * mm, y, f"{unit_price:,.2f} {currency} × {qty}")
                y -= line_h * 0.95

        # ── Summary (subtotal / VAT / service / discount) ───────────────
        if summary:
            y = self._dashed_rule(c, margin, W - margin, y)
            for label, val in summary:
                self._draw_mixed(c, margin, y, label + ":", 6.8, color=gray, align="left")
                self._draw_mixed(c, W - margin, y, str(val), 6.8, color=colors.HexColor("#374151"), align="right")
                y -= line_h * 0.85

        # ── Total ─────────────────────────────────────────────────────────
        y -= line_h * 0.3
        c.setStrokeColor(primary)
        c.setLineWidth(1.2)
        c.line(margin, y, W - margin, y)
        y -= line_h * 1.3

        self._draw_mixed(c, margin, y, "الإجمالي", 9.5, bold=True, color=primary, align="left")
        c.setFont(FONT_LATIN_BOLD, 14)
        c.setFillColor(accent)
        c.drawRightString(W - margin, y - 1, f"{total:,.2f} {currency}")
        y -= line_h * 1.9

        if note:
            self._draw_mixed(c, W / 2, y, note, 6.5, color=gray, align="center")
            y -= line_h

        if qr_data:
            try:
                import qrcode as _qrcode
                qr_img = _qrcode.make(qr_data)
                qr_buf = BytesIO()
                qr_img.save(qr_buf, format="PNG")
                qr_buf.seek(0)
                from reportlab.lib.utils import ImageReader
                c.drawImage(ImageReader(qr_buf), (W - qr_size) / 2, y - qr_size, qr_size, qr_size)
            except ImportError:
                pass

        self._add_footer(c, W, footer or self.app_name, accent)
        c.save()
        return buf.getvalue()

    # ── Excel: Multi-sheet Workbook ─────────────────────────────────────

    def excel(
        self,
        sheets: list[dict],
        *,
        title: str = "",
        freeze_rows: bool = True,
        auto_width: bool = True,
    ) -> bytes:
        """Generate a multi-sheet Excel workbook."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise RuntimeError("pip install openpyxl to use Excel generation")

        _thin = Side(style="thin", color=_BORDER_COLOR.lstrip("#"))
        _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

        primary = self.primary_color
        accent = self.accent_color
        light = "F8F9FA"
        white = "FFFFFF"

        def _hcell(ws, row, col, value, bg=primary, fg=white, bold=True, width=None):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = Font(bold=bold, color=fg, size=10, name="Calibri")
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = _border
            if width:
                ws.column_dimensions[get_column_letter(col)].width = width
            return cell

        def _dcell(ws, row, col, value, bold=False, align="left", num_fmt=None, bg=None):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = Font(bold=bold, size=10, name="Calibri")
            cell.alignment = Alignment(horizontal=align, vertical="center")
            cell.border = _border
            if bg:
                cell.fill = PatternFill("solid", fgColor=bg)
            if num_fmt:
                cell.number_format = num_fmt
            return cell

        wb = Workbook()
        if wb.active is not None:
            wb.remove(wb.active)  # remove default empty sheet

        gen_str = datetime.now().strftime("%Y-%m-%d %H:%M")

        for si, sheet_def in enumerate(sheets):
            ws = wb.create_sheet(title=sheet_def.get("name", f"Sheet{si+1}"))
            ws.sheet_view.showGridLines = False

            s_headers = sheet_def.get("headers", [])
            s_rows = sheet_def.get("rows", [])
            s_summary = sheet_def.get("summary", {})
            s_col_types = sheet_def.get("col_types", [])
            n_cols = len(s_headers)

            ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
            t = ws["A1"]
            t.value = f"{self.app_name}  —  {sheet_def.get('name', title)}  |  {gen_str}"
            t.font = Font(bold=True, size=13, color=white, name="Calibri")
            t.fill = PatternFill("solid", fgColor=primary)
            t.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 28

            for ci, hdr in enumerate(s_headers, 1):
                _hcell(ws, 2, ci, hdr, bg=accent, fg=primary)
            ws.row_dimensions[2].height = 22

            if freeze_rows:
                ws.freeze_panes = "A3"

            for ri, row in enumerate(s_rows, 3):
                bg = light if (ri - 3) % 2 == 0 else None
                for ci, val in enumerate(row, 1):
                    col_type = s_col_types[ci - 1] if ci - 1 < len(s_col_types) else "text"
                    num_fmt = None
                    align = "left"
                    if col_type == "currency":
                        num_fmt = "#,##0.00"
                        align = "right"
                    elif col_type == "number":
                        align = "right"
                    elif col_type == "percent":
                        num_fmt = "0.0%"
                        align = "right"
                    _dcell(ws, ri, ci, val, align=align, num_fmt=num_fmt, bg=bg)

            if s_summary:
                total_rows = len(s_rows)
                sr = 3 + total_rows
                keys = list(s_summary.keys())
                vals = list(s_summary.values())
                merge_end = max(1, n_cols - len(vals))
                if merge_end > 1:
                    ws.merge_cells(f"A{sr}:{get_column_letter(merge_end)}{sr}")
                lbl_val = keys[0] if len(keys) == 1 else "الإجمالي / Total"
                lbl = ws.cell(row=sr, column=1, value=lbl_val)
                lbl.font = Font(bold=True, size=10, color=white, name="Calibri")
                lbl.fill = PatternFill("solid", fgColor=accent)
                lbl.alignment = Alignment(horizontal="left", vertical="center")
                lbl.border = _border
                for vi, (_, v) in enumerate(s_summary.items()):
                    ci = merge_end + vi + 1
                    if ci > n_cols:
                        break
                    cell = ws.cell(row=sr, column=ci, value=v)
                    cell.font = Font(bold=True, size=10, color=primary, name="Calibri")
                    cell.fill = PatternFill("solid", fgColor=accent)
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.border = _border
                    if isinstance(v, (int, float)):
                        cell.number_format = "#,##0.00"

            if auto_width:
                for col_cells in ws.columns:
                    max_len = 0
                    col_letter = get_column_letter(col_cells[0].column)
                    for cell in col_cells:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ── Helpers ───────────────────────────────────────────────────────────

    def _t(self, text: str) -> str:
        """Apply Arabic reshaping + bidi if available and rtl=True."""
        if not self.rtl or not text:
            return text
        try:
            import arabic_reshaper
            from bidi.algorithm import get_display
            result = get_display(arabic_reshaper.reshape(text))
            return result if isinstance(result, str) else result.decode("utf-8")
        except ImportError:
            return text

    def _draw_mixed(
        self, c, x: float, y: float, text: str, size: float, *,
        bold: bool = False, color=None, align: str = "left",
    ) -> None:
        """يرسم سطر ممكن يحتوي عربي وإنجليزي مع بعض، كل جزء بالخط اللي
        فعليًا بيغطي حروفه (راجع تعليق _split_script_runs فوق للسبب).
        ``text`` نص منطقي خام (بيتم reshape+bidi هنا داخليًا) — الـ callers
        متعمّلوش self._t() بأيديهم قبل ما يبعتوه هنا."""
        if not text:
            return
        shaped = self._t(text) if self.rtl else text
        runs = _split_script_runs(shaped)
        if not runs:
            return

        def font_for(is_ar: bool) -> str:
            if is_ar:
                return FONT_AR_BOLD if bold else FONT_AR
            return FONT_LATIN_BOLD if bold else FONT_LATIN

        total_w = sum(c.stringWidth(r, font_for(a), size) for r, a in runs)
        if align == "right":
            cx = x - total_w
        elif align == "center":
            cx = x - total_w / 2
        else:
            cx = x

        if color is not None:
            c.setFillColor(color)
        for run, is_ar in runs:
            font = font_for(is_ar)
            c.setFont(font, size)
            c.drawString(cx, y, run)
            cx += c.stringWidth(run, font, size)

    @staticmethod
    def _dashed_rule(c, x1: float, x2: float, y: float) -> float:
        """رسم فاصل متقطع (نمط إيصالات حرارية قياسي)، ويرجّع الـy بعد الفاصل."""
        from reportlab.lib import colors
        c.setStrokeColor(colors.HexColor(_BORDER_COLOR))
        c.setLineWidth(0.6)
        c.setDash(2, 2)
        c.line(x1, y, x2, y)
        c.setDash()
        return y - 3.6

    def _add_footer(self, c, W, text: str, accent_color) -> None:
        from reportlab.lib import colors
        c.setFillColor(colors.HexColor("#1A1A2E"))
        c.rect(0, 0, W, 35, fill=True, stroke=False)
        self._draw_mixed(c, W / 2, 20, text, 8, color=accent_color, align="center")
        c.setFillColor(colors.HexColor("#AAAAAA"))
        c.setFont(FONT_LATIN, 7)
        c.drawCentredString(W / 2, 10, datetime.now().strftime("%Y-%m-%d %H:%M"))
